"""Leitura, validacao e exportacao de planilhas de baixas SERASA."""

import os
import re
from datetime import datetime
from typing import Dict, Iterable, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


COLUNA_CPF_CNPJ_NORMALIZADO = "_CPF_CNPJ_NORMALIZADO"
COLUNA_AUTO_NORMALIZADO = "_AUTO_INFRACAO_NORMALIZADO"
COLUNA_LINHA_ORIGINAL = "_LINHA_ORIGINAL"

COLUNAS_RESULTADO = [
    "STATUS_BAIXA",
    "MENSAGEM_PORTAL",
    "TENTATIVAS",
    "DATA_PROCESSAMENTO",
    "OBSERVACAO",
]

SINONIMOS_CPF_CNPJ = (
    "cpf/cnpj",
    "cpf cnpj",
    "cpf_cnpj",
    "cpf",
    "cnpj",
    "documento",
    "contribuinte",
)

SINONIMOS_AUTO = (
    "auto de infracao",
    "auto de infração",
    "auto infracao",
    "auto infração",
    "identificador do debito",
    "identificador do débito",
    "identificador",
    "numero do auto",
    "número do auto",
    "auto",
)


def normalizar_nome_coluna(nome) -> str:
    texto = str(nome).strip().lower()
    texto = texto.replace("_", " ").replace("-", " ")
    texto = re.sub(r"\s+", " ", texto)
    return texto


def normalizar_cpf_cnpj(valor) -> Optional[str]:
    if pd.isna(valor):
        return None
    digitos = re.sub(r"\D", "", str(valor).strip())
    if not digitos:
        return None
    if len(digitos) > 14:
        digitos = digitos[-14:]
    return digitos


def formatar_cpf_cnpj(valor) -> str:
    digitos = normalizar_cpf_cnpj(valor)
    if not digitos:
        return ""
    if len(digitos) == 11:
        return f"{digitos[:3]}.{digitos[3:6]}.{digitos[6:9]}-{digitos[9:]}"
    if len(digitos) == 14:
        return f"{digitos[:2]}.{digitos[2:5]}.{digitos[5:8]}/{digitos[8:12]}-{digitos[12:]}"
    return digitos


def normalizar_auto(valor) -> Optional[str]:
    if pd.isna(valor):
        return None
    texto = str(valor).strip()
    if not texto or texto.lower() in {"nan", "none"}:
        return None
    return re.sub(r"\s+", "", texto).upper()


def _detectar_coluna(colunas: Iterable[str], sinonimos: Iterable[str]) -> Optional[str]:
    mapa = {normalizar_nome_coluna(coluna): coluna for coluna in colunas}
    for sinonimo in sinonimos:
        chave = normalizar_nome_coluna(sinonimo)
        if chave in mapa:
            return mapa[chave]

    for chave, original in mapa.items():
        if any(normalizar_nome_coluna(sinonimo) in chave for sinonimo in sinonimos):
            return original
    return None


def detectar_colunas(df: pd.DataFrame) -> Dict[str, str]:
    coluna_cpf_cnpj = _detectar_coluna(df.columns, SINONIMOS_CPF_CNPJ)
    coluna_auto = _detectar_coluna(df.columns, SINONIMOS_AUTO)
    faltantes = []
    if not coluna_cpf_cnpj:
        faltantes.append("CPF/CNPJ")
    if not coluna_auto:
        faltantes.append("Auto de Infracao")
    if faltantes:
        raise ValueError(
            "Colunas obrigatorias nao encontradas: "
            + ", ".join(faltantes)
            + ". Informe uma planilha com CPF/CNPJ e Auto de Infracao."
        )
    return {"cpf_cnpj": coluna_cpf_cnpj, "auto_infracao": coluna_auto}


def ler_planilha(caminho: str) -> pd.DataFrame:
    if not os.path.isfile(caminho):
        raise FileNotFoundError(f"Arquivo nao encontrado: {caminho}")

    extensao = os.path.splitext(caminho)[1].lower()
    if extensao in {".xlsx", ".xlsm"}:
        df = pd.read_excel(caminho, dtype=str, engine="openpyxl")
    elif extensao == ".xls":
        df = pd.read_excel(caminho, dtype=str, engine="xlrd")
    elif extensao == ".csv":
        df = pd.read_csv(caminho, dtype=str, sep=None, engine="python", encoding="utf-8-sig")
    else:
        raise ValueError("Formato nao suportado. Use .xlsx, .xlsm, .xls ou .csv.")

    df.columns = df.columns.astype(str).str.strip()
    df = df.dropna(how="all").copy()
    if df.empty:
        raise ValueError("A planilha esta vazia.")

    df[COLUNA_LINHA_ORIGINAL] = range(2, len(df) + 2)
    return df


def preparar_base(caminho: str) -> tuple[pd.DataFrame, Dict[str, str]]:
    df = ler_planilha(caminho)
    colunas = detectar_colunas(df)

    df[COLUNA_CPF_CNPJ_NORMALIZADO] = df[colunas["cpf_cnpj"]].apply(normalizar_cpf_cnpj)
    df[COLUNA_AUTO_NORMALIZADO] = df[colunas["auto_infracao"]].apply(normalizar_auto)

    for coluna in COLUNAS_RESULTADO:
        if coluna not in df.columns:
            df[coluna] = ""

    sem_cpf = df[COLUNA_CPF_CNPJ_NORMALIZADO].isna()
    sem_auto = df[COLUNA_AUTO_NORMALIZADO].isna()
    df.loc[sem_cpf | sem_auto, "STATUS_BAIXA"] = "IGNORADO"
    df.loc[sem_cpf, "OBSERVACAO"] = "CPF/CNPJ vazio ou invalido"
    df.loc[sem_auto, "OBSERVACAO"] = df.loc[sem_auto, "OBSERVACAO"].apply(
        lambda atual: "Auto de Infracao vazio ou invalido"
        if not atual
        else f"{atual}; Auto de Infracao vazio ou invalido"
    )

    return df, colunas


def aplicar_resultado(
    df: pd.DataFrame,
    indice,
    status: str,
    mensagem: str,
    tentativas: int,
    observacao: str = "",
) -> None:
    df.at[indice, "STATUS_BAIXA"] = status
    df.at[indice, "MENSAGEM_PORTAL"] = mensagem or ""
    df.at[indice, "TENTATIVAS"] = tentativas
    df.at[indice, "DATA_PROCESSAMENTO"] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    df.at[indice, "OBSERVACAO"] = observacao or ""


def exportar_resultado(df: pd.DataFrame, caminho_saida: str) -> None:
    pasta_saida = os.path.dirname(caminho_saida)
    if pasta_saida:
        os.makedirs(pasta_saida, exist_ok=True)
    colunas_ocultas = {
        COLUNA_CPF_CNPJ_NORMALIZADO,
        COLUNA_AUTO_NORMALIZADO,
        COLUNA_LINHA_ORIGINAL,
    }
    export_df = df[[coluna for coluna in df.columns if coluna not in colunas_ocultas]].copy()
    export_df.to_excel(caminho_saida, index=False, engine="openpyxl")
    formatar_excel(caminho_saida)


def formatar_excel(caminho: str) -> None:
    wb = load_workbook(caminho)
    ws = wb.active
    ws.title = "Resultado Baixas"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    status_colors = {
        "SUCESSO": "C6EFCE",
        "ERRO": "FFC7CE",
        "IGNORADO": "FFEB9C",
        "PENDENTE_MAPEAMENTO": "D9EAF7",
        "PARADO": "E7E6E6",
        "SEM_DIVIDA": "FFEB9C",
    }
    headers = {cell.value: cell.column for cell in ws[1]}
    status_col = headers.get("STATUS_BAIXA")

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        if status_col:
            status = ws.cell(row=row[0].row, column=status_col).value
            fill = status_colors.get(str(status or "").upper())
            if fill:
                for cell in row:
                    cell.fill = PatternFill("solid", fgColor=fill)

    for column_cells in ws.columns:
        max_len = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            max_len = max(max_len, len(str(cell.value or "")))
        ws.column_dimensions[column_letter].width = min(max(max_len + 2, 12), 60)

    if ws.max_row >= 2 and ws.max_column >= 1:
        ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        tabela = Table(displayName="TabelaResultadoBaixas", ref=ref)
        estilo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        tabela.tableStyleInfo = estilo
        ws.add_table(tabela)

    ws.freeze_panes = "A2"
    wb.save(caminho)
